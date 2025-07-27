from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from collections import defaultdict, Counter
import random
from collections import defaultdict
from django.db.models import Max
from django.utils import timezone
from .models import Student, ScreenSettings, ExamResult, STATUS_CHOICES
from .forms import StudentForm, ScreenSettingsForm
from django.contrib import messages
from django.http import HttpResponse
from django.utils.encoding import smart_str
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime, time, timedelta
from django.utils.timezone import make_aware, localtime, get_current_timezone
from collections import defaultdict
from django.db.models import Max
from django.contrib.admin.views.decorators import staff_member_required
from .models import Student, ExamResult, ScreenSettings
from django.shortcuts import render
import os
import csv
import openpyxl
from django.conf import settings
import json
from django.db import models


def get_room_count():
    settings = ScreenSettings.objects.last()
    return settings.room_count if settings else 5

def get_least_loaded_room():
    room_count = get_room_count()
    rooms = list(range(1, room_count + 1))
    counts = Counter(Student.objects.values_list('room', flat=True))
    for r in rooms:
        counts.setdefault(r, 0)
    return min(rooms, key=lambda r: counts[r])

@login_required
@staff_member_required
def public_screen(request):
    
    # Order by room and student position instead of number
    students = list(Student.objects.all().order_by('room', 'position'))

    # Fetch the latest ExamResult for each student
    latest_ids = (
        ExamResult.objects
        .values('number')
        .annotate(latest_id=Max('id'))
        .values_list('latest_id', flat=True)
    )
    results = ExamResult.objects.filter(id__in=latest_ids)

    # Map student number → grade/result
    latest_results = {
        res.number: {
            'grade': res.grade,
            'result': res.result
        }
        for res in results
    }

    # Attach latest grade/result to student objects
    for student in students:
        result = latest_results.get(student.number)
        student.latest_grade = result['grade'] if result else None
        student.latest_result = result['result'] if result else None

    # Group students by room
    students_by_room = defaultdict(list)
    for student in students:
        students_by_room[student.room].append(student)

    # Optional: sort within room (finished last, but keep order otherwise)
    for room, room_students in students_by_room.items():
        room_students.sort(key=lambda s: (s.status == 'finished', s.position))

    # Load screen settings
    screen_settings = ScreenSettings.get_settings()
    estimate_minutes = screen_settings.estimate_time_per_student or 5

    # Get exam start time from settings (datetime.time) or default to 7:00 AM
    exam_start_time_value = screen_settings.exam_start_time or time(7, 0)

    # Combine with today's date to form a naive datetime
    today = datetime.today()
    naive_start_datetime = datetime.combine(today, exam_start_time_value)

    # Make timezone aware
    tz = get_current_timezone()
    exam_start_time = make_aware(naive_start_datetime, timezone=tz)
    exam_start_time = localtime(exam_start_time)  # Convert to local time if needed


    student_wait_times = {}
    for room, students_in_room in students_by_room.items():
        waiting_students = [s for s in students_in_room if s.status in ['waiting', 'on_waiting_list']]
        for idx, student in enumerate(waiting_students):
            wait_delta = timedelta(minutes=(idx + 1) * estimate_minutes)
            wait_time = exam_start_time + wait_delta
            student_wait_times[student.number] = wait_time.strftime("%I:%M").lstrip("0")


    return render(request, 'screen/public_screen.html', {
        'students_by_room': students_by_room,
        'rooms': range(1, get_room_count() + 1),
        'student_wait_times': student_wait_times,
    })
@login_required
def clear_students(request):
    if request.method == 'POST':
        Student.objects.all().delete()
    return redirect('/screen/add-student')

def apply_automatic_status_for_room(room):
    students = Student.objects.filter(room=room).exclude(status='finished').order_by('position')

    for i, student in enumerate(students):
        if i == 0:
            student.status = 'in_exam'
        elif 1 <= i <= 6:
            student.status = 'waiting'
        else:
            student.status = 'on_waiting_list'
        student.save()


def update_student_status(request, student_number):
    if request.method == 'POST':
        student = get_object_or_404(Student, number=student_number)
        new_status = request.POST.get('status')

        if new_status == "remove":
            student.delete()
            messages.success(request, f"تم حذف الطالب {student.name}")

        elif new_status.startswith("move:"):
            try:
                new_room = int(new_status.split(":")[1])
                student.room = new_room
                # Assign next available position in new room
                max_position = (
                    Student.objects.filter(room=new_room)
                    .aggregate(max_pos=models.Max('position'))['max_pos']
                )
                student.position = (max_position or 0) + 1
                student.save()
                messages.success(request, f"تم نقل الطالب {student.name} إلى لجنة {new_room}")
            except:
                messages.error(request, "حدث خطأ أثناء نقل الطالب")

        else:
            student.status = new_status
            student.save()
            messages.success(request, f"تم تحديث حالة الطالب {student.name} إلى {student.get_status_display()}")

        # Always reapply automatic status for the two rooms involved
        apply_automatic_status_for_room(student.room)
    
    return redirect('add_student')



def apply_automatic_status():
    settings = ScreenSettings.objects.last()
    waiting_limit = settings.waiting_count if settings else 5

    rooms = Student.objects.values_list('room', flat=True).distinct()

    for room in rooms:
        # ORDER BY 'position' to reflect current order after moves
        students = Student.objects.filter(room=room).exclude(status='finished').order_by('position')
        for i, student in enumerate(students):
            if i == 0:
                student.status = 'in_exam'
            elif 1 <= i <= waiting_limit:
                student.status = 'waiting'
            else:
                student.status = 'on_waiting_list'
            student.save()



@login_required
def trigger_automatic_status(request):
    apply_automatic_status()
    return redirect('/screen/add-student')
@login_required
@staff_member_required
def add_student(request):
    form = StudentForm()

    if request.method == 'POST' and 'name' in request.POST:
        form = StudentForm(request.POST)
        if form.is_valid():
            student = form.save(commit=False)

            # Set position to the next available one in the room
            max_position = (
                Student.objects.filter(room=student.room)
                .aggregate(max_pos=Max('position'))
            )['max_pos']
            student.position = (max_position or 0) + 1
            student.save()
            messages.success(request, 'تمت إضافة الطالب بنجاح.')
            return redirect('add_student')
        else:
            messages.error(request, 'الرجاء التحقق من صحة البيانات.')

    rooms = list(range(1, get_room_count() + 1))
    students_by_room = {
        room: Student.objects.filter(room=room).order_by('position')
        for room in rooms
    }

    all_institutes = (
        Student.objects.values_list('institute_name', flat=True)
        .distinct()
        .order_by('institute_name')
    )

    context = {
        'form': form,
        'rooms': rooms,
        'students_by_room': students_by_room,
        'status_choices': STATUS_CHOICES,
        'all_institutes': all_institutes,
    }

    return render(request, 'screen/add_student.html', context)




@require_POST
@login_required
def submit_grade(request, student_number, subroom):
    student = get_object_or_404(Student, number=student_number)
    
    # Parse exam config
    if student.exam_type == "gh":
        expected_questions = 3
        pass_threshold = 80
    elif student.exam_type == "nz":
        expected_questions = 5
        pass_threshold = 90
    else:
        expected_questions = 0
        pass_threshold = 0

    # Get submitted final grade for this subroom from form
    try:
        final_grade = float(request.POST.get("final_grade", "100"))
    except ValueError:
        final_grade = 100

    # Save/update this subroom's partial result with a temporary 'result' value
    exam_result, created = ExamResult.objects.update_or_create(
        number=student.number,
        sub_room=subroom,
        defaults={
            'name': student.name,
            'grade': final_grade,
            'result': 'قيد التقييم',  # temporary placeholder to satisfy NOT NULL
            'room': student.room,
        }
    )

    # Check all subroom grades for this student
    subroom_results = ExamResult.objects.filter(number=student.number).exclude(sub_room=0)
    grades = [er.grade for er in subroom_results]

    if len(grades) >= 2:  # assuming 2 subrooms always
        avg_grade = sum(grades) / len(grades)
        result = "ناجح" if avg_grade >= pass_threshold else "إعادة"

        # Save final summary with sub_room=0 to indicate final average
        ExamResult.objects.update_or_create(
            number=student.number,
            sub_room=0,
            defaults={
                'name': student.name,
                'grade': avg_grade,
                'result': result,
                'room': student.room,
            }
        )

        # Mark student finished
        student.status = 'finished'
        student.save()

        # Log to CSV after final grade calculated
        csv_path = os.path.join(settings.BASE_DIR, 'grades_log.csv')
        file_exists = os.path.isfile(csv_path)
        with open(csv_path, mode='a', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(['Student Number', 'Name', 'Final Grade', 'Result', 'Sub Room'])
            writer.writerow([student.number, student.name, avg_grade, result, 'final_average'])

        # Redirect to room page (choose your preferred subroom here)
        return redirect(f"/mobileapp/room/room{student.room}/1/")

    else:
        avg_grade = sum(grades) / len(grades)
        messages.info(request, f"تم تسجيل درجة هذا القسم. الدرجة الحالية هي {avg_grade:.2f}. انتظر القسم الآخر لإكمال التقييم.")
        # Redirect to parent room page even on first submission
        return redirect(f"/mobileapp/room/room{student.room}/1/")

@login_required
@staff_member_required
def edit_settings(request):
    old_settings = ScreenSettings.get_settings()
    old_room_count = old_settings.room_count if old_settings else 0

    if request.method == 'POST':
        form = ScreenSettingsForm(request.POST, instance=old_settings)
        if form.is_valid():
            new_settings = form.save(commit=False)
            new_room_count = new_settings.room_count

            if new_room_count != old_room_count:
                # Reassign all students to fit new room count
                students = list(Student.objects.all().exclude(status="finished"))
                room_positions = {room: 0 for room in range(1, new_room_count + 1)}

                for student in students:
                    min_room = min(room_positions, key=room_positions.get)
                    room_positions[min_room] += 1
                    student.room = min_room
                    student.position = room_positions[min_room]
                    student.save()

            new_settings.save()

            apply_automatic_status()

            return redirect('/screen/add-student')

    else:
        form = ScreenSettingsForm(instance=old_settings)

    return render(request, 'screen/edit_settings.html', {'form': form})


EXAM_TYPE_MAP = {
    "غيباً": "gh",
    "نظراً": "nz",
}

@login_required
@require_POST
def upload_excel(request):
    file = request.FILES.get("file")
    if not file or not file.name.endswith(".xlsx"):
        return redirect('/screen/add-student')

    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    current_number = Student.objects.aggregate(Max('number'))['number__max'] or 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        number_cell = str(row[0]).strip() if row[0] else ''
        name = str(row[1]).strip() if row[1] else ''

        if not name or "الاسم" in name:
            continue

        try:
            father_name = str(row[2]).strip() if row[2] else ''
            birth_year = int(row[3]) if row[3] else None
            institute = str(row[4]).strip() if row[4] else ''
            exam_type_arabic = str(row[5]).strip() if row[5] else ''
            exam_type = EXAM_TYPE_MAP.get(exam_type_arabic, None)
            parts = int(row[6]) if row[6] else 0

            if not exam_type:
                print(f"Unknown exam type: {exam_type_arabic} in row: {row}")
                continue

            current_number += 1

            Student.objects.create(
                number=current_number,
                name=name,
                father_name=father_name,
                birth_date=f"{birth_year}-01-01" if birth_year else None,
                institute_name=institute,
                exam_type=exam_type,
                memorized_parts=parts
            )
        except Exception as e:
            print("Error in row:", row, str(e))
            continue

    apply_automatic_status()
    return redirect('/screen/add-student')

def remove_student(request, student_number):
    if request.method == 'POST':
        student = get_object_or_404(Student, number=student_number)
        student.delete()
        messages.success(request, "تم حذف الطالب بنجاح.")
    return redirect('add_student')  

@require_POST
@login_required
def clear_all_results(request):
    # Delete all ExamResult entries
    ExamResult.objects.all().delete()

    # Delete only students with status "finished"
    Student.objects.filter(status="finished").delete()

    # Redirect to the add-student page after clearing
    return redirect('/screen/add-student')

@login_required
def export_students_excel(request):
    institute_name = request.GET.get('institute')
    if not institute_name:
        return HttpResponse("يرجى اختيار اسم المعهد", status=400)

    settings = ScreenSettings.get_settings()
    estimated_time_per_student = settings.estimate_time_per_student or 5
    exam_start_time = settings.exam_start_time or datetime.time(8, 0)
    start_minutes = exam_start_time.hour * 60 + exam_start_time.minute

    # Build room-wise queues
    students = Student.objects.all().order_by('room', 'number')
    room_queues = defaultdict(list)
    for student in students:
        room_queues[student.room].append(student)

    # Assign estimated times based on position in room queue
    student_times = {}
    max_queue_length = max(len(queue) for queue in room_queues.values())
    for index in range(max_queue_length):
        for room in sorted(room_queues.keys()):
            queue = room_queues[room]
            if index < len(queue):
                student = queue[index]
                est_minutes = start_minutes + index * estimated_time_per_student
                est_time = f"{est_minutes // 60:02}:{est_minutes % 60:02}"
                student_times[student.number] = est_time

    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.sheet_view.rightToLeft = True
    current_row = 1

    def write_headers():
        nonlocal current_row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        ws.cell(row=current_row, column=1, value="الجمهورية العربية السورية").alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=1).font = Font(size=14, bold=True)
        current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        ws.cell(row=current_row, column=1, value="استمارة  اختبار الأجزاء المتفرقة").alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=1).font = Font(size=12, bold=True)
        current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        ws.cell(row=current_row, column=1, value="وزارة الأوقاف").alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=8)
        ws.cell(row=current_row, column=5, value="مركز الحسنين").alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=5).font = Font(bold=True)
        current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        ws.cell(row=current_row, column=1, value="مديرية معاهد تحفيظ القرآن الكريم بدمشق").alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 2

    def write_table_header():
        nonlocal current_row
        headers = ['الرقم', 'الاسم والكنية', 'اسم الأب', 'تاريخ الولادة', 'اسم المعهد', 'غيباً/نظراً', 'الأجزاء المحفوظة', 'الوقت المتوقع']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=i, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        current_row += 1

    def write_student_row(student):
        nonlocal current_row
        exam_type = student.get_exam_type_display() if hasattr(student, 'get_exam_type_display') else student.exam_type
        est_time = student_times.get(student.number, '')
        ws.append([
            student.number,
            smart_str(student.name),
            smart_str(student.father_name or ''),
            student.birth_date.year if student.birth_date else '',
            smart_str(student.institute_name or ''),
            exam_type,
            student.memorized_parts or '',
            est_time
        ])
        current_row += 1

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    write_headers()
    write_table_header()

    if institute_name == '__all__':
        grouped = defaultdict(list)
        for s in students:
            grouped[s.institute_name].append(s)

        for institute in sorted(grouped.keys()):
            sorted_students = sorted(grouped[institute], key=lambda x: student_times.get(x.number, ''))
            for stu in sorted_students:
                write_student_row(stu)
        filename = f"All_Students_{datetime.today().date()}.xlsx"
    else:
        filtered_students = Student.objects.filter(institute_name=institute_name).order_by('room', 'number')
        for student in sorted(filtered_students, key=lambda x: student_times.get(x.number, '')):
            write_student_row(student)
        filename = f"Students_{institute_name}_{datetime.today().date()}.xlsx"


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@require_POST
@login_required
def move_student_position(request, number):
    direction = request.POST.get("direction")
    student = get_object_or_404(Student, number=number)

    same_room_students = list(Student.objects.filter(room=student.room).order_by('position'))

    try:
        index = same_room_students.index(student)
    except ValueError:
        return redirect('add_student')

    if direction == "up" and index > 0:
        same_room_students[index], same_room_students[index - 1] = same_room_students[index - 1], same_room_students[index]
    elif direction == "down" and index < len(same_room_students) - 1:
        same_room_students[index], same_room_students[index + 1] = same_room_students[index + 1], same_room_students[index]

    for idx, s in enumerate(same_room_students):
        s.position = idx
        s.save()

    apply_automatic_status()  # reassign statuses based on new order

    return redirect('add_student')
